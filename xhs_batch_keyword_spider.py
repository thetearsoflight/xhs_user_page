from xhs_keyword_spider_v2 import XHSKeywordSpiderV2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import time
import random
import argparse
import json
import sqlite3


PROGRESS_FILE = 'data/batch_progress.json'
DB_DIR = 'db'
DB_FILE = os.path.join(DB_DIR, 'notes_history.db')


def init_db():
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notes (
            note_id TEXT PRIMARY KEY,
            title TEXT,
            nickname TEXT,
            keyword TEXT,
            note_type TEXT,
            likes TEXT,
            likes_num INTEGER,
            note_url TEXT,
            crawl_time REAL
        )
    ''')
    conn.commit()
    return conn


def get_existing_note_ids(conn):
    cursor = conn.cursor()
    cursor.execute('SELECT note_id FROM notes')
    return set(row[0] for row in cursor.fetchall())


def save_notes_to_db(conn, notes, keyword):
    cursor = conn.cursor()
    crawl_time = time.time()
    for note in notes:
        cursor.execute('''
            INSERT OR IGNORE INTO notes 
            (note_id, title, nickname, keyword, note_type, likes, likes_num, note_url, crawl_time)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            note.get('note_id', ''),
            note.get('title', ''),
            note.get('nickname', ''),
            keyword,
            note.get('note_type', ''),
            note.get('likes', '0'),
            note.get('likes_num', 0),
            note.get('note_url', ''),
            crawl_time
        ))
    conn.commit()


def load_keywords(filename='keywords.txt'):
    keywords = []
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    keywords.append(line)
    except FileNotFoundError:
        print(f"关键词文件 {filename} 不存在，请先创建")
    return keywords


def load_progress():
    if not os.path.exists(PROGRESS_FILE):
        return {'completed_keywords': [], 'all_notes': []}
    try:
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {'completed_keywords': [], 'all_notes': []}


def save_progress(progress):
    os.makedirs(os.path.dirname(PROGRESS_FILE), exist_ok=True)
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def clear_progress():
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)
        print("已清除进度文件")


def deduplicate_by_title(notes):
    seen_titles = set()
    unique_notes = []
    dup_count = 0
    for note in notes:
        normalized = note.get('title', '').strip()
        if not normalized:
            unique_notes.append(note)
            continue
        if normalized not in seen_titles:
            seen_titles.add(normalized)
            unique_notes.append(note)
        else:
            dup_count += 1
    return unique_notes, dup_count


def save_merged_excel(notes, like_threshold, output_dir='data'):
    filtered = [n for n in notes if n.get('likes_num', 0) > like_threshold]
    print(f"\n筛选: {len(notes)} 篇 → 点赞>{like_threshold}: {len(filtered)} 篇")

    if not filtered:
        print("没有达标笔记，不生成汇总文件")
        return None

    unique_notes, dup_count = deduplicate_by_title(filtered)
    print(f"去重: {len(filtered)} 篇 → 删除{dup_count}条重复标题 → {len(unique_notes)} 篇")

    if not unique_notes:
        print("去重后无笔记，不生成汇总文件")
        return None

    os.makedirs(output_dir, exist_ok=True)
    filename = os.path.join(output_dir, f'关键词汇总_{time.strftime("%Y%m%d_%H%M%S")}.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总数据"

    headers = ['序号', '标题', '作者', '搜索关键词', '类型', '点赞数', '详情页URL']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    link_font = Font(color='0563C1', underline='single')

    for row, note in enumerate(unique_notes, 2):
        ws.cell(row=row, column=1, value=row - 1).border = thin_border
        ws.cell(row=row, column=2, value=note.get('title', '')).border = thin_border
        ws.cell(row=row, column=3, value=note.get('nickname', '')).border = thin_border
        ws.cell(row=row, column=4, value=note.get('keyword', '')).border = thin_border

        note_type = note.get('note_type', '')
        type_display = '视频' if note_type == 'video' else '图文'
        ws.cell(row=row, column=5, value=type_display).border = thin_border

        ws.cell(row=row, column=6, value=note.get('likes', '0')).border = thin_border

        url_cell = ws.cell(row=row, column=7)
        url_value = note.get('note_url', '')
        url_cell.value = url_value
        url_cell.hyperlink = url_value
        url_cell.font = link_font
        url_cell.border = thin_border

        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 55

    ws.freeze_panes = 'A2'

    wb.save(filename)
    print(f"汇总数据已保存到: {filename}")
    return filename


def main():
    print("=" * 60)
    print("小红书批量关键词爬虫（接口监听版）")
    print("=" * 60)
    print()

    parser = argparse.ArgumentParser(description='小红书批量关键词爬虫')
    parser.add_argument('-f', '--file', type=str, default='resources/keywords.txt', help='关键词库文件路径，默认resources/keywords.txt')
    parser.add_argument('-n', '--num', type=int, default=30, help='每个关键词采集的达标笔记数量，默认30篇')
    parser.add_argument('-l', '--likes', type=int, default=200, help='点赞数阈值，默认200')
    parser.add_argument('--no-sort-time', action='store_true', help='不按最新排序（默认按最新排序）')
    parser.add_argument('--restart', action='store_true', help='忽略进度文件，从头开始')
    parser.add_argument('--timeout', type=int, default=300, help='单个关键词超时秒数，默认300秒')
    parser.add_argument('--clear-db', action='store_true', help='清空历史笔记数据库，从头开始去重')
    args = parser.parse_args()

    keywords = load_keywords(args.file)
    if not keywords:
        print(f"关键词库为空，请检查 {args.file}")
        return

    target_count = args.num if args.num > 0 else 50
    sort_by_time = not args.no_sort_time

    progress = load_progress()
    if args.restart:
        clear_progress()
        progress = {'completed_keywords': [], 'all_notes': []}

    completed_set = set(progress['completed_keywords'])
    all_notes = progress['all_notes']

    remaining = [kw for kw in keywords if kw not in completed_set]

    if completed_set:
        print(f"发现进度文件: 已完成 {len(completed_set)}/{len(keywords)} 个关键词")
        if remaining:
            print(f"剩余待处理: {len(remaining)} 个关键词")
        else:
            print("所有关键词已爬取完毕，直接生成汇总文件")
            if all_notes:
                save_merged_excel(all_notes, args.likes)
            return
    else:
        print(f"从 {args.file} 加载了 {len(keywords)} 个关键词")
        for i, kw in enumerate(keywords, 1):
            print(f"  {i}. {kw}")

    print()
    print(f"每个关键词目标: {target_count} 篇点赞>{args.likes}的笔记")
    print(f"排序: {'最新' if sort_by_time else '综合'}")
    print(f"单关键词超时: {args.timeout}秒")
    print("注意: 请确保已登录小红书账号")
    print()

    spider = XHSKeywordSpiderV2()
    spider.like_threshold = args.likes
    results = []

    # 初始化数据库，加载历史笔记ID用于去重
    conn = init_db()
    if args.clear_db:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM notes')
        conn.commit()
        print("已清空历史笔记数据库")
    existing_ids = get_existing_note_ids(conn)
    print(f"历史笔记库已加载: {len(existing_ids)} 篇")

    try:
        for i, keyword in enumerate(remaining, 1):
            print(f"\n{'=' * 60}")
            print(f"正在处理第 {i}/{len(remaining)} 个关键词: {keyword}")
            print(f"{'=' * 60}")

            if i > 1:
                gap = random.uniform(60, 120)
                print(f"关键词间隔等待 {gap:.1f} 秒（防封禁）...")
                time.sleep(gap)

            spider.notes_data = []
            spider.seen_ids = set()
            spider.keyword = keyword

            crawl_success = False
            crawl_start = time.time()

            try:
                notes = spider.crawl_keyword_notes(keyword, target_count=target_count, sort_by_time=sort_by_time, max_scrolls=25)
                elapsed = time.time() - crawl_start

                if elapsed > args.timeout:
                    print(f"⚠ 关键词 '{keyword}' 超时({elapsed:.0f}秒>{args.timeout}秒)，跳过")

                for note in notes:
                    note['keyword'] = keyword

                # 去重：只保留本次新爬取的笔记
                new_notes = [n for n in notes if n['note_id'] not in existing_ids]
                dup_count = len(notes) - len(new_notes)
                if dup_count > 0:
                    print(f"去重: 过滤掉 {dup_count} 篇历史笔记，本次新增 {len(new_notes)} 篇")

                # 只保留点赞数达标的笔记
                qualified_notes = [n for n in new_notes if n.get('likes_num', 0) > args.likes]
                unqualified_count = len(new_notes) - len(qualified_notes)
                if unqualified_count > 0:
                    print(f"筛选: 过滤掉 {unqualified_count} 篇点赞<={args.likes}的笔记，不存入数据库")

                # 将达标笔记存入数据库
                if qualified_notes:
                    save_notes_to_db(conn, qualified_notes, keyword)
                    existing_ids.update(n['note_id'] for n in qualified_notes)

                all_notes.extend(qualified_notes)

                qualified = len(qualified_notes)
                results.append({
                    'keyword': keyword,
                    'total': len(notes),
                    'new': len(new_notes),
                    'qualified': qualified,
                    'success': True,
                })
                crawl_success = True
                print(f"\n关键词 '{keyword}' 完成: 总计{len(notes)}篇，新增{len(new_notes)}篇，达标{qualified}篇，耗时{elapsed:.0f}秒")

            except Exception as e:
                print(f"关键词 '{keyword}' 爬取出错: {e}")
                import traceback
                traceback.print_exc()
                results.append({
                    'keyword': keyword,
                    'total': 0,
                    'qualified': 0,
                    'success': False,
                })

            if crawl_success or keyword not in completed_set:
                progress['completed_keywords'].append(keyword)
                progress['all_notes'] = all_notes
                save_progress(progress)
                print(f"进度已保存 (已完成 {len(progress['completed_keywords'])}/{len(keywords)})")

        print(f"\n{'=' * 60}")
        print("所有关键词爬取完毕")
        print(f"{'=' * 60}")

        print(f"\n各关键词统计:")
        total_all = 0
        new_all = 0
        qualified_all = 0
        for r in results:
            status = "✓" if r['success'] else "✗"
            print(f"  [{status}] {r['keyword']:<15} | 总计: {r['total']:>4} | 新增: {r.get('new', 0):>4} | 达标: {r['qualified']:>4}")
            total_all += r['total']
            new_all += r.get('new', 0)
            qualified_all += r['qualified']
        print(f"  {'─' * 60}")
        print(f"  {'本次合计':<15} | 总计: {total_all:>4} | 新增: {new_all:>4} | 达标: {qualified_all:>4}")
        print(f"  {'累计合计':<15} | 总计: {len(all_notes):>4}")

        if all_notes:
            excel_file = save_merged_excel(all_notes, args.likes)

            print()
            print("=" * 60)
            print("批量爬取完成！")
            print(f"关键词数: {len(keywords)}")
            print(f"总笔记数: {len(all_notes)}")
            if excel_file:
                print(f"汇总文件: {excel_file}")
            print("=" * 60)

            clear_progress()
        else:
            print("\n未获取到任何笔记数据")

        conn.close()
        spider.close()

    except KeyboardInterrupt:
        print(f"\n\n用户中断！进度已保存，下次运行将从断点继续")
        print(f"已完成: {len(progress['completed_keywords'])}/{len(keywords)} 个关键词")
        conn.close()
        spider.close()

    except Exception as e:
        print(f"批量爬取过程中出错: {e}")
        import traceback
        traceback.print_exc()
        print(f"进度已保存，下次运行将从断点继续")
        conn.close()
        input("\n按回车键关闭浏览器...")
        spider.close()


if __name__ == '__main__':
    main()
