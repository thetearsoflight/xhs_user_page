from DrissionPage import ChromiumPage
import json
import time
import re
import os
from urllib.parse import urlencode, parse_qs, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class XHSSpider:
    def __init__(self):
        self.page = ChromiumPage()
        self.notes_data = []
        self.user_name = ''

    def parse_user_id_from_url(self, url):
        """从小红书主页URL中提取用户ID"""
        patterns = [
            r'/user/profile/(\w+)',
            r'/user/profile/([^?/]+)',
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return None

    def get_user_name(self):
        """获取博主用户名"""
        try:
            name_selectors = [
                'css:.user-name',
                'css:.user-nickname',
                'css:[class*="userName"]',
                'css:[class*="nickname"]',
                'css:.user-info .name',
                'css:.profile-name',
            ]
            for selector in name_selectors:
                try:
                    name_elem = self.page.ele(selector, timeout=1)
                    if name_elem:
                        name = name_elem.text.strip()
                        if name:
                            self.user_name = name
                            print(f"获取到博主用户名: {name}")
                            return name
                except:
                    continue

            js_code = """
            const nameElem = document.querySelector('.user-name, .user-nickname, [class*="userName"], [class*="nickname"]');
            return nameElem ? nameElem.innerText.trim() : '';
            """
            name = self.page.run_js(js_code)
            if name:
                self.user_name = name
                print(f"获取到博主用户名: {name}")
                return name
        except Exception as e:
            print(f"获取用户名失败: {e}")

        self.user_name = 'unknown_user'
        return self.user_name

    def scroll_page(self, scroll_times=20, scroll_pause=3):
        """滚动页面加载更多笔记"""
        print(f"开始滚动页面，计划滚动 {scroll_times} 次...")
        for i in range(scroll_times):
            self.page.scroll.down(400)
            time.sleep(scroll_pause)
            print(f"第 {i + 1}/{scroll_times} 次滚动完成")

    def _extract_notes_from_page(self):
        """从当前页面提取笔记"""
        note_selectors = [
            'css:.note-item',
            'css:[class*="note-item"]',
            'css:.feeds-page .note-item',
            'css:.user-page .note-item',
            'css:section[class*="note"] > div',
            'css:.waterfall-item',
        ]

        for selector in note_selectors:
            try:
                elements = self.page.eles(selector, timeout=2)
                if elements and len(elements) > 0:
                    for note_elem in elements:
                        try:
                            note_info = self.extract_note_info(note_elem)
                            if note_info and note_info['note_id']:
                                if not any(n['note_id'] == note_info['note_id'] for n in self.notes_data):
                                    self.notes_data.append(note_info)
                        except:
                            continue
                    break
            except:
                continue

    def scroll_and_extract(self, scroll_times=20, scroll_pause=3):
        """滚动页面并逐步提取笔记"""
        print(f"开始滚动并提取，计划滚动 {scroll_times} 次...")
        for i in range(scroll_times):
            self.page.scroll.down(400)
            time.sleep(scroll_pause)

            # 每滚动2次提取一次笔记
            if (i + 1) % 2 == 0:
                self._extract_notes_from_page()
                print(f"第 {i + 1}/{scroll_times} 次滚动完成，当前已提取 {len(self.notes_data)} 篇笔记")
            else:
                print(f"第 {i + 1}/{scroll_times} 次滚动完成")

    def extract_note_info(self, note_element):
        """从笔记元素中提取信息"""
        try:
            # 获取笔记链接
            link_elem = note_element.ele('css:a[href*="/explore/"]', timeout=0.5)
            if not link_elem:
                return None

            note_url = link_elem.attr('href')
            if not note_url.startswith('http'):
                note_url = 'https://www.xiaohongshu.com' + note_url

            # 获取笔记ID
            note_id_match = re.search(r'/explore/(\w+)', note_url)
            note_id = note_id_match.group(1) if note_id_match else ''

            # 获取标题
            title = ''
            try:
                title_elem = note_element.ele('css:.title, .desc, span[class*="title"]', timeout=0.5)
                if title_elem:
                    title = title_elem.text.strip()
            except:
                pass

            # 如果没有标题，尝试获取描述
            if not title:
                try:
                    desc_elem = note_element.ele('css:.desc span, .content span', timeout=0.5)
                    if desc_elem:
                        title = desc_elem.text.strip()[:50]  # 截取前50字
                except:
                    pass

            # 获取点赞数
            likes = '0'
            try:
                # 尝试多种点赞数选择器
                like_selectors = [
                    'css:.like-wrapper .count',
                    'css:.likes .count',
                    'css:span[class*="like"]',
                    'css:.interaction span',
                    'css:.count',
                    'css:[class*="like"] span',
                    'css:.info span',
                    'css:.meta span',
                    'css:span',
                ]
                for selector in like_selectors:
                    try:
                        like_elem = note_element.ele(selector, timeout=0.3)
                        if like_elem:
                            likes_text = like_elem.text.strip()
                            # 检查是否为数字格式（包括万、w、k等）
                            if likes_text and any(c.isdigit() for c in likes_text):
                                likes = likes_text
                                break
                    except:
                        continue
            except Exception as e:
                print(f"获取点赞数出错: {e}")
                pass

            # 获取封面图
            cover_image = ''
            try:
                img_elem = note_element.ele('css:img[class*="img"], css:.cover img', timeout=0.3)
                if img_elem:
                    cover_image = img_elem.attr('src') or img_elem.attr('data-src')
            except:
                pass

            return {
                'note_id': note_id,
                'title': title,
                'likes': likes,
                'note_url': note_url,
                'cover_image': cover_image,
            }
        except Exception as e:
            print(f"提取笔记信息时出错: {e}")
            return None

    def crawl_user_notes(self, user_url, scroll_times=10):
        """爬取用户主页的所有笔记"""
        print(f"正在访问用户主页: {user_url}")

        self.page.get(user_url)
        time.sleep(3)

        self.get_user_name()

        # 先提取页面已有的笔记（前几条）
        print("正在提取初始笔记...")
        self._extract_notes_from_page()

        # 滚动并持续提取
        self.scroll_and_extract(scroll_times=scroll_times)

        # 最后再提取一次确保没有遗漏
        self._extract_notes_from_page()

        print(f"\n共提取到 {len(self.notes_data)} 篇笔记")

        # 打印所有笔记的点赞数（用于调试）
        print("\n提取到的笔记数据预览：")
        for i, note in enumerate(self.notes_data[:10], 1):
            parsed_likes = self.parse_likes(note['likes'])
            print(f"  {i}. 点赞: {note['likes']:>8} -> 解析: {parsed_likes:>6} | 标题: {note['title'][:30]}...")
        if len(self.notes_data) > 10:
            print(f"  ... 还有 {len(self.notes_data) - 10} 篇笔记")

        return self.notes_data

    def parse_likes(self, likes_str):
        """解析点赞数为数字"""
        if not likes_str:
            return 0
        likes_str = str(likes_str).strip()
        try:
            if '万' in likes_str:
                num = float(likes_str.replace('万', ''))
                return int(num * 10000)
            elif 'w' in likes_str.lower():
                num = float(likes_str.lower().replace('w', ''))
                return int(num * 10000)
            elif 'k' in likes_str.lower():
                num = float(likes_str.lower().replace('k', ''))
                return int(num * 1000)
            else:
                return int(float(likes_str))
        except:
            return 0

    def save_to_excel(self):
        """保存数据到Excel文件"""
        filtered_notes = [
            note for note in self.notes_data
            if self.parse_likes(note.get('likes', '0')) > 90
        ]

        print(f"原始笔记数: {len(self.notes_data)}, 筛选后(点赞>90): {len(filtered_notes)}")

        if not filtered_notes:
            print("没有点赞数大于90的笔记，不生成Excel文件")
            return None

        filename = f"{self.user_name}_notes.xlsx" if self.user_name else 'xhs_notes.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "笔记数据"

        headers = ['序号', '标题', '点赞数', '详情页URL']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        link_font = Font(color='0563C1', underline='single')

        for row, note in enumerate(filtered_notes, 2):
            ws.cell(row=row, column=1, value=row - 1).border = thin_border
            ws.cell(row=row, column=2, value=note.get('title', '')).border = thin_border
            ws.cell(row=row, column=3, value=note.get('likes', '0')).border = thin_border

            url_cell = ws.cell(row=row, column=4)
            url_value = note.get('note_url', '')
            url_cell.value = url_value
            url_cell.hyperlink = url_value
            url_cell.font = link_font
            url_cell.border = thin_border

            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 60

        ws.freeze_panes = 'A2'

        wb.save(filename)
        print(f"数据已保存到: {filename}")
        return filename

    def close(self):
        """关闭浏览器"""
        self.page.quit()
        print("浏览器已关闭")


def main():
    print("=" * 60)
    print("小红书博主笔记爬虫")
    print("=" * 60)
    print()

    # 获取用户输入
    user_url = input("请输入小红书博主主页URL: ").strip()

    if not user_url:
        print("URL不能为空！")
        return

    # 验证URL格式
    if 'xiaohongshu.com' not in user_url:
        print("请输入有效的小红书URL！")
        return

    # 设置滚动次数
    scroll_input = input("请输入滚动次数(默认10次，每次加载约10-20篇笔记): ").strip()
    scroll_times = int(scroll_input) if scroll_input.isdigit() else 10

    print()
    print("正在启动爬虫...")
    print("注意: 请确保已登录小红书账号，否则可能无法获取完整数据")
    print()

    # 创建爬虫实例
    spider = XHSSpider()

    try:
        # 爬取笔记
        notes = spider.crawl_user_notes(user_url, scroll_times=scroll_times)

        if notes:
            excel_file = spider.save_to_excel()

            print()
            print("=" * 60)
            print("爬取完成！")
            print(f"博主: {spider.user_name}")
            print(f"共获取 {len(notes)} 篇笔记")
            print(f"数据已保存到: {excel_file}")
            print("=" * 60)
        else:
            print("未获取到任何笔记数据，请检查：")
            print("1. URL是否正确")
            print("2. 是否已登录小红书")
            print("3. 页面是否正常加载")

    except Exception as e:
        print(f"爬取过程中出错: {e}")
        import traceback
        traceback.print_exc()

    finally:
        input("\n按回车键关闭浏览器...")
        spider.close()


if __name__ == '__main__':
    main()
