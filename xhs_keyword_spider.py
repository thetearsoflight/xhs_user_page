from DrissionPage import ChromiumPage
import json
import time
import re
import os
from urllib.parse import urlencode, parse_qs, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class XHSKeywordSpider:
    def __init__(self):
        self.page = ChromiumPage()
        self.notes_data = []
        self.keyword = ''

    def search_by_keyword(self, keyword):
        """根据关键词搜索笔记"""
        self.keyword = keyword
        search_url = f"https://www.xiaohongshu.com/search_result?keyword={keyword}&source=web_explore_feed"
        print(f"正在搜索关键词: {keyword}")
        print(f"访问URL: {search_url}")
        
        self.page.get(search_url)
        
        import random

        # 模拟人类浏览行为 - 页面加载后的随机等待
        page_load_delay = random.uniform(2, 4)
        print(f"等待页面加载中... ({page_load_delay:.1f}秒)")
        time.sleep(page_load_delay)

        # 随机延迟（模拟人类查看页面）
        if random.random() < 0.5:
            for _ in range(random.randint(2, 5)):
                time.sleep(random.uniform(0.1, 0.3))
        
        # 点击筛选按钮并选择最新
        print("正在进行筛选操作...")
        try:
            # 参考my_xhs_ai.py的筛选方法，使用XPath选择器
            print("使用XPath选择器查找筛选按钮...")
            try:
                # 定位筛选按钮
                shaixuan = self.page.ele('xpath://div[@class="search-layout__top"]//span[text()="筛选"]', timeout=5)
                if shaixuan:
                    print("找到筛选按钮")
                    # 悬浮到筛选按钮上
                    shaixuan.hover()
                    print("已悬浮到筛选按钮")
                    time.sleep(2)  # 等待悬浮框出现
                    
                    # 定位最新选项
                    latest_button = self.page.ele('xpath://div[@class="search-layout__top"]//span[text()="最新"]', timeout=3)
                    if latest_button:
                        print("找到'最新'选项")
                        # 点击最新选项
                        latest_button.click()
                        print("已点击'最新'选项")
                        # 等待页面刷新
                        time.sleep(4)
                        print("已选择'最新'筛选选项并等待页面刷新")
                    else:
                        print("未找到'最新'选项")
                else:
                    print("未找到筛选按钮")
            except Exception as e:
                print(f"XPath选择器操作失败: {e}")
                
                # 备选方案：尝试使用CSS选择器
                print("尝试使用CSS选择器...")
                try:
                    # 尝试多种筛选按钮选择器
                    filter_selectors = [
                        'css:.filter',
                        'css:.sort',
                        'css:.filter-btn',
                        'css:.sort-btn',
                        'css:[class*="filter"]',
                        'css:[class*="sort"]',
                        'css:button:contains(筛选)',
                        'css:button:contains(排序)',
                        'css:div[class*="filter"]',
                        'css:div[class*="sort"]',
                    ]
                    
                    filter_button = None
                    for selector in filter_selectors:
                        try:
                            button = self.page.ele(selector, timeout=2)
                            if button:
                                filter_button = button
                                print(f"找到筛选按钮: {selector}")
                                break
                        except:
                            continue
                    
                    if filter_button:
                        filter_button.hover()
                        print("已悬浮到筛选按钮")
                        time.sleep(3)
                        
                        # 尝试多种最新选项选择器
                        latest_selectors = [
                            'css:.filter-item:contains(最新)',
                            'css:.sort-item:contains(最新)',
                            'css:.option:contains(最新)',
                            'css:li:contains(最新)',
                            'css:div:contains(最新)',
                            'css:span:contains(最新)',
                        ]
                        
                        latest_option = None
                        for selector in latest_selectors:
                            try:
                                option = self.page.ele(selector, timeout=3)
                                if option:
                                    latest_option = option
                                    print(f"找到'最新'选项: {selector}")
                                    break
                            except:
                                continue
                        
                        if latest_option:
                            latest_option.click()
                            print("已点击'最新'选项")
                            time.sleep(4)
                            print("已选择'最新'筛选选项并等待页面刷新")
                        else:
                            print("未找到'最新'选项")
                    else:
                        print("未找到筛选按钮")
                except Exception as e:
                    print(f"CSS选择器操作也失败: {e}")
        except Exception as e:
            print(f"筛选操作失败: {e}")
            pass

    def scroll_page(self, scroll_times=20, scroll_pause=3):
        """滚动页面加载更多笔记"""
        print(f"开始滚动页面，计划滚动 {scroll_times} 次...")
        for i in range(scroll_times):
            self.page.scroll.down(400)
            time.sleep(scroll_pause)
            print(f"第 {i + 1}/{scroll_times} 次滚动完成")

    def _extract_notes_from_page(self):
        """从当前页面提取笔记"""
        note_selectors = [
            'css:.note-item',
            'css:[class*="note-item"]',
            'css:.feeds-page .note-item',
            'css:section[class*="note"] > div',
            'css:.waterfall-item',
        ]

        for selector in note_selectors:
            try:
                elements = self.page.eles(selector, timeout=2)
                if elements and len(elements) > 0:
                    for note_elem in elements:
                        try:
                            note_info = self.extract_note_info(note_elem)
                            if note_info and note_info['note_id']:
                                if not any(n['note_id'] == note_info['note_id'] for n in self.notes_data):
                                    self.notes_data.append(note_info)
                        except:
                            continue
                    break
            except:
                continue

    def scroll_and_extract(self, scroll_times=20, scroll_pause=3):
        """滚动页面并逐步提取笔记"""
        print(f"开始滚动并提取，计划滚动 {scroll_times} 次...")
        for i in range(scroll_times):
            self.page.scroll.down(400)
            time.sleep(scroll_pause)

            # 每滚动2次提取一次笔记
            if (i + 1) % 2 == 0:
                self._extract_notes_from_page()
                print(f"第 {i + 1}/{scroll_times} 次滚动完成，当前已提取 {len(self.notes_data)} 篇笔记")
            else:
                print(f"第 {i + 1}/{scroll_times} 次滚动完成")

    def extract_note_info(self, note_element):
        """从笔记元素中提取信息"""
        try:
            # 获取笔记链接
            link_elem = note_element.ele('css:a[href*="/explore/"]', timeout=0.5)
            if not link_elem:
                return None

            note_url = link_elem.attr('href')
            if not note_url.startswith('http'):
                note_url = 'https://www.xiaohongshu.com' + note_url

            # 获取笔记ID
            note_id_match = re.search(r'/explore/(\w+)', note_url)
            note_id = note_id_match.group(1) if note_id_match else ''

            # 获取标题
            title = ''
            try:
                title_elem = note_element.ele('css:.title, .desc, span[class*="title"]', timeout=0.5)
                if title_elem:
                    title = title_elem.text.strip()
            except:
                pass

            # 如果没有标题，尝试获取描述
            if not title:
                try:
                    desc_elem = note_element.ele('css:.desc span, .content span', timeout=0.5)
                    if desc_elem:
                        title = desc_elem.text.strip()[:50]  # 截取前50字
                except:
                    pass

            # 获取点赞数
            likes = '0'
            try:
                # 尝试多种点赞数选择器
                like_selectors = [
                    'css:.like-wrapper .count',
                    'css:.likes .count',
                    'css:span[class*="like"]',
                    'css:.interaction span',
                    'css:.count',
                    'css:[class*="like"] span',
                    'css:.info span',
                    'css:.meta span',
                    'css:span',
                ]
                for selector in like_selectors:
                    try:
                        like_elem = note_element.ele(selector, timeout=0.3)
                        if like_elem:
                            likes_text = like_elem.text.strip()
                            # 检查是否为数字格式（包括万、w、k等）
                            if likes_text and any(c.isdigit() for c in likes_text):
                                likes = likes_text
                                break
                    except:
                        continue
            except Exception as e:
                print(f"获取点赞数出错: {e}")
                pass

            # 获取封面图
            cover_image = ''
            try:
                img_elem = note_element.ele('css:img[class*="img"], css:.cover img', timeout=0.3)
                if img_elem:
                    cover_image = img_elem.attr('src') or img_elem.attr('data-src')
            except:
                pass

            return {
                'note_id': note_id,
                'title': title,
                'likes': likes,
                'note_url': note_url,
                'cover_image': cover_image,
            }
        except Exception as e:
            print(f"提取笔记信息时出错: {e}")
            return None

    def count_qualified_notes(self):
        """统计点赞>90的笔记数量"""
        return sum(1 for note in self.notes_data if self.parse_likes(note.get('likes', '0')) > 90)

    def crawl_keyword_notes(self, keyword, target_count=50):
        """根据关键词爬取笔记，直到达到目标数量"""
        self.search_by_keyword(keyword)

        # 先提取页面已有的笔记（前几条）
        print("正在提取初始笔记...")
        self._extract_notes_from_page()

        qualified_count = self.count_qualified_notes()
        print(f"当前达标笔记数: {qualified_count}/{target_count}")

        # 如果还没达到目标，继续滚动提取
        max_scrolls = 50  # 最大滚动次数，防止无限循环
        no_new_count = 0  # 连续没有新笔记的次数

        import random

        for scroll_idx in range(max_scrolls):
            if qualified_count >= target_count:
                print(f"\n已达到目标数量 {target_count} 篇达标笔记，停止爬取")
                break

            # 模拟人类滚动行为 - 随机滚动距离和速度
            scroll_distance = random.randint(300, 600)  # 随机滚动距离
            scroll_pause = random.uniform(1.5, 3.5)  # 随机暂停时间
            
            # 滚动页面
            self.page.scroll.down(scroll_distance)
            
            # 随机添加一些人类行为
            if random.random() < 0.3:  # 30%概率
                # 随机延迟（模拟人类查看页面）
                time.sleep(random.uniform(0.2, 0.5))

            time.sleep(scroll_pause)

            # 提取笔记
            prev_count = len(self.notes_data)
            self._extract_notes_from_page()
            new_notes = len(self.notes_data) - prev_count

            qualified_count = self.count_qualified_notes()

            if new_notes == 0:
                no_new_count += 1
                print(f"第 {scroll_idx + 1} 次滚动: 无新笔记 (连续{no_new_count}次)")
                # 连续3次没有新笔记，认为已经到底
                if no_new_count >= 3:
                    print("\n已连续3次没有新笔记，认为已到达页面底部")
                    break
            else:
                no_new_count = 0
                print(f"第 {scroll_idx + 1} 次滚动: 新增{new_notes}篇，达标{qualified_count}/{target_count}")

        if qualified_count < target_count:
            print(f"\n搜索结果已爬取完毕，达标笔记仅{qualified_count}篇（目标{target_count}篇）")

        print(f"\n共提取到 {len(self.notes_data)} 篇笔记，其中达标{qualified_count}篇")

        # 打印所有笔记的点赞数（用于调试）
        print("\n提取到的笔记数据预览（前10篇）：")
        for i, note in enumerate(self.notes_data[:10], 1):
            parsed_likes = self.parse_likes(note['likes'])
            qualified = "✓" if parsed_likes > 90 else "✗"
            print(f"  {i}. [{qualified}] 点赞: {note['likes']:>8} -> {parsed_likes:>6} | {note['title'][:25]}...")
        if len(self.notes_data) > 10:
            print(f"  ... 还有 {len(self.notes_data) - 10} 篇笔记")

        return self.notes_data

    def parse_likes(self, likes_str):
        """解析点赞数为数字"""
        if not likes_str:
            return 0
        likes_str = str(likes_str).strip()
        try:
            if '万' in likes_str:
                num = float(likes_str.replace('万', ''))
                return int(num * 10000)
            elif 'w' in likes_str.lower():
                num = float(likes_str.lower().replace('w', ''))
                return int(num * 10000)
            elif 'k' in likes_str.lower():
                num = float(likes_str.lower().replace('k', ''))
                return int(num * 1000)
            else:
                return int(float(likes_str))
        except:
            return 0

    def save_to_excel(self):
        """保存数据到Excel文件"""
        filtered_notes = [
            note for note in self.notes_data
            if self.parse_likes(note.get('likes', '0')) > 90
        ]

        print(f"原始笔记数: {len(self.notes_data)}, 筛选后(点赞>90): {len(filtered_notes)}")

        if not filtered_notes:
            print("没有点赞数大于90的笔记，不生成Excel文件")
            return None

        # 创建data目录（如果不存在）
        data_dir = 'data'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
            print(f"创建目录: {data_dir}/")

        filename = os.path.join(data_dir, f"{self.keyword}_notes.xlsx" if self.keyword else 'xhs_notes.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = "笔记数据"

        headers = ['序号', '标题', '点赞数', '详情页URL']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        link_font = Font(color='0563C1', underline='single')

        for row, note in enumerate(filtered_notes, 2):
            ws.cell(row=row, column=1, value=row - 1).border = thin_border
            ws.cell(row=row, column=2, value=note.get('title', '')).border = thin_border
            ws.cell(row=row, column=3, value=note.get('likes', '0')).border = thin_border

            url_cell = ws.cell(row=row, column=4)
            url_value = note.get('note_url', '')
            url_cell.value = url_value
            url_cell.hyperlink = url_value
            url_cell.font = link_font
            url_cell.border = thin_border

            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 60

        ws.freeze_panes = 'A2'

        wb.save(filename)
        print(f"数据已保存到: {filename}")
        return filename

    def close(self):
        """关闭浏览器"""
        self.page.quit()
        print("浏览器已关闭")


def main():
    print("=" * 60)
    print("小红书关键词搜索爬虫")
    print("=" * 60)
    print()

    # 获取用户输入
    keyword = input("请输入搜索关键词: ").strip()
    
    if not keyword:
        print("关键词不能为空！")
        return

    target_input = input("请输入需要采集的达标笔记数量(点赞>90，默认50篇): ").strip()
    target_count = int(target_input) if target_input.isdigit() else 50

    print()
    print("正在启动爬虫...")
    print(f"目标: 采集 {target_count} 篇点赞>90的笔记")
    print("注意: 请确保已登录小红书账号，否则可能无法获取完整数据")
    print()

    spider = XHSKeywordSpider()

    try:
        notes = spider.crawl_keyword_notes(keyword, target_count=target_count)

        if notes:
            excel_file = spider.save_to_excel()

            print()
            print("=" * 60)
            print("爬取完成！")
            print(f"关键词: {keyword}")
            print(f"共获取 {len(notes)} 篇笔记")
            print(f"数据已保存到: {excel_file}")
            print("=" * 60)
        else:
            print("未获取到任何笔记数据")

    except Exception as e:
        print(f"爬取过程中出错: {e}")
        import traceback
        traceback.print_exc()

    finally:
        input("\n按回车键关闭浏览器...")
        spider.close()


if __name__ == '__main__':
    main()