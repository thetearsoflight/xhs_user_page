from xhs_user_spider import XHSSpider
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import time
import random
import argparse
import json


PROGRESS_FILE = 'data/batch_user_progress.json'


def load_urls(filename='resources/urls.txt'):
    urls = []
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    urls.append(line)
    except FileNotFoundError:
        print(f"URL文件 {filename} 不存在，请先创建")
    return urls


def load_progress():
    if not os.path.exists(PROGRESS_FILE):
        return {'completed_urls': [], 'all_results': []}
    try:
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {'completed_urls': [], 'all_results': []}


def save_progress(progress):
    os.makedirs(os.path.dirname(PROGRESS_FILE), exist_ok=True)
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def clear_progress():
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)
        print("已清除进度文件")


def save_monitor_excel(all_qualified_notes, like_threshold, output_dir='data'):
    if not all_qualified_notes:
        print("没有达标笔记，不生成汇总文件")
        return None

    os.makedirs(output_dir, exist_ok=True)
    filename = os.path.join(output_dir, f'监控_{time.strftime("%Y%m%d_%H%M%S")}.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "监控数据"

    headers = ['序号', '博主', '标题', '点赞数', '详情页URL']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    link_font = Font(color='0563C1', underline='single')

    for row, note in enumerate(all_qualified_notes, 2):
        ws.cell(row=row, column=1, value=row - 1).border = thin_border
        ws.cell(row=row, column=2, value=note.get('nickname', '')).border = thin_border
        ws.cell(row=row, column=3, value=note.get('title', '')).border = thin_border
        ws.cell(row=row, column=4, value=note.get('likes', '0')).border = thin_border

        url_cell = ws.cell(row=row, column=5)
        url_value = note.get('note_url', '')
        url_cell.value = url_value
        url_cell.hyperlink = url_value
        url_cell.font = link_font
        url_cell.border = thin_border

        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 60

    ws.freeze_panes = 'A2'

    wb.save(filename)
    print(f"监控汇总数据已保存到: {filename}")
    return filename


def crawl_single_blogger_check(spider, url, check_count, like_threshold, blogger_index, total_count):
    print("\n" + "=" * 60)
    print(f"正在处理第 {blogger_index}/{total_count} 个博主")
    print(f"URL: {url}")
    print(f"模式: 检查前{check_count}篇笔记，点赞>{like_threshold}")
    print("=" * 60)

    try:
        spider.page.get(url)
        
        page_load_delay = random.uniform(2, 4)
        print(f"等待页面加载中... ({page_load_delay:.1f}秒)")
        time.sleep(page_load_delay)

        _simulate_human_behavior(spider.page, duration=3)

        spider.get_user_name()

        print("正在提取笔记...")
        spider._extract_notes_from_page()

        max_scrolls = 30
        no_new_count = 0
        scroll_count = 0

        for scroll_idx in range(max_scrolls):
            if len(spider.notes_data) >= check_count:
                print(f"\n已提取{len(spider.notes_data)}篇笔记，达到检查上限{check_count}篇")
                break

            scroll_distance = random.randint(300, 600)
            scroll_pause = random.uniform(1.5, 3.5)
            
            spider.page.scroll.down(scroll_distance)
            scroll_count += 1

            _simulate_human_behavior(spider.page, duration=2)

            time.sleep(scroll_pause)

            if (scroll_idx + 1) % 2 == 0:
                prev_count = len(spider.notes_data)
                spider._extract_notes_from_page()
                new_notes = len(spider.notes_data) - prev_count

                if new_notes == 0:
                    no_new_count += 1
                    print(f"第 {scroll_idx + 1} 次滚动: 无新笔记 (连续{no_new_count}次)")
                    if no_new_count >= 3:
                        print("\n已连续3次没有新笔记，认为已到达页面底部")
                        break
                else:
                    no_new_count = 0
                    print(f"第 {scroll_idx + 1} 次滚动: 新增{new_notes}篇，当前共{len(spider.notes_data)}篇")

            if scroll_count % 5 == 0:
                print(f"请求频率控制: 休息3秒...")
                time.sleep(3)

        notes_data = spider.notes_data[:check_count]
        
        qualified_notes = [note for note in notes_data if spider.parse_likes(note.get('likes', '0')) > like_threshold]
        qualified_count = len(qualified_notes)

        for note in qualified_notes:
            note['nickname'] = spider.user_name

        print()
        print("-" * 60)
        print(f"博主 {spider.user_name} 检查完成")
        print(f"检查笔记: {len(notes_data)} 篇")
        print(f"达标笔记(点赞>{like_threshold}): {qualified_count} 篇")
        
        if qualified_count > 0:
            print(f"\n达标笔记列表:")
            for i, note in enumerate(qualified_notes, 1):
                print(f"  {i}. [{note['likes']}] {note['title'][:40]}... | {note['note_url']}")
        else:
            print("无达标笔记")
        
        print("-" * 60)

        return {
            'url': url,
            'name': spider.user_name,
            'total': len(notes_data),
            'qualified': qualified_count,
            'qualified_notes': qualified_notes,
            'file': None,
            'success': True
        }

    except Exception as e:
        print(f"处理博主时出错: {e}")
        import traceback
        traceback.print_exc()
        return {
            'url': url,
            'name': '错误',
            'total': 0,
            'qualified': 0,
            'file': None,
            'success': False
        }


def _simulate_human_behavior(page, duration=5):
    start_time = time.time()
    actions = 0
    
    while time.time() - start_time < duration:
        action_type = random.choice(['scroll_small', 'hover', 'move_mouse', 'pause'])
        
        if action_type == 'scroll_small':
            scroll_amount = random.randint(-100, 100)
            if scroll_amount > 0:
                page.scroll.down(abs(scroll_amount))
            else:
                page.scroll.up(abs(scroll_amount))
            time.sleep(random.uniform(0.5, 1.5))
            
        elif action_type == 'hover':
            try:
                elements = page.eles('css:a', timeout=1)
                if elements:
                    elem = random.choice(elements[:10])
                    elem.hover()
                    time.sleep(random.uniform(0.3, 0.8))
            except:
                pass
                
        elif action_type == 'move_mouse':
            x = random.randint(200, 800)
            y = random.randint(200, 600)
            try:
                page.run_js(f'window.dispatchEvent(new MouseEvent("mousemove", {{clientX: {x}, clientY: {y}}}))')
            except:
                pass
            time.sleep(random.uniform(0.2, 0.5))
            
        else:
            time.sleep(random.uniform(0.5, 1.5))
        
        actions += 1
        
        if random.random() < 0.3:
            break
    
    return actions


def crawl_single_blogger(spider, url, target_count, blogger_index, total_count):
    print("\n" + "=" * 60)
    print(f"正在处理第 {blogger_index}/{total_count} 个博主")
    print(f"URL: {url}")
    print("=" * 60)

    try:
        notes = spider.crawl_user_notes(url, target_count=target_count)

        if notes:
            excel_file = spider.save_to_excel()
            qualified_count = spider.count_qualified_notes()

            print()
            print("-" * 60)
            print(f"博主 {spider.user_name} 处理完成")
            print(f"总笔记: {len(notes)} 篇，达标: {qualified_count} 篇")
            if excel_file:
                print(f"保存文件: {excel_file}")
            print("-" * 60)

            return {
                'url': url,
                'name': spider.user_name,
                'total': len(notes),
                'qualified': qualified_count,
                'file': excel_file,
                'success': True
            }
        else:
            print(f"未从该博主获取到笔记数据")
            return {
                'url': url,
                'name': '未知',
                'total': 0,
                'qualified': 0,
                'file': None,
                'success': False
            }

    except Exception as e:
        print(f"处理博主时出错: {e}")
        import traceback
        traceback.print_exc()
        return {
            'url': url,
            'name': '错误',
            'total': 0,
            'qualified': 0,
            'file': None,
            'success': False
        }


def generate_summary_report(results, target_count, check_mode=False, check_count=0):
    print("\n" + "=" * 60)
    print("批量爬取汇总报告")
    print("=" * 60)

    total_bloggers = len(results)
    success_bloggers = sum(1 for r in results if r['success'])
    total_notes = sum(r['total'] for r in results)
    total_qualified = sum(r['qualified'] for r in results)

    print(f"\n总体统计:")
    print(f"  博主总数: {total_bloggers}")
    print(f"  成功爬取: {success_bloggers}")
    print(f"  失败/无数据: {total_bloggers - success_bloggers}")
    print(f"  总笔记数: {total_notes}")
    print(f"  总达标数: {total_qualified}")
    
    if check_mode:
        bloggers_with_qualified = sum(1 for r in results if r['qualified'] > 0)
        print(f"  有达标笔记的博主: {bloggers_with_qualified}/{total_bloggers}")

    print(f"\n各博主详情:")
    for i, result in enumerate(results, 1):
        status = "✓" if result['success'] else "✗"
        print(f"  {i}. [{status}] {result['name'][:15]:<15} | 达标: {result['qualified']:>3} | 总笔记: {result['total']:>3}")

    data_dir = 'data'
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)

    mode_str = f'check{check_count}_' if check_mode else ''
    report_file = os.path.join(data_dir, f'batch_user_report_{mode_str}{time.strftime("%Y%m%d_%H%M%S")}.txt')
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("批量爬取汇总报告\n")
        f.write("=" * 60 + "\n\n")
        if check_mode:
            f.write(f"监控模式: 检查每个博主前{check_count}篇笔记\n\n")
        f.write(f"总体统计:\n")
        f.write(f"  博主总数: {total_bloggers}\n")
        f.write(f"  成功爬取: {success_bloggers}\n")
        f.write(f"  失败/无数据: {total_bloggers - success_bloggers}\n")
        f.write(f"  总笔记数: {total_notes}\n")
        f.write(f"  总达标数: {total_qualified}\n")
        if check_mode:
            bloggers_with_qualified = sum(1 for r in results if r['qualified'] > 0)
            f.write(f"  有达标笔记的博主: {bloggers_with_qualified}/{total_bloggers}\n")
        f.write("\n各博主详情:\n")
        for i, result in enumerate(results, 1):
            status = "成功" if result['success'] else "失败"
            f.write(f"  {i}. [{status}] {result['name']}\n")
            f.write(f"      URL: {result['url']}\n")
            f.write(f"      达标: {result['qualified']}, 总笔记: {result['total']}\n")
            if result['file']:
                f.write(f"      文件: {result['file']}\n")
            f.write("\n")

    print(f"\n汇总报告已保存到: {report_file}")
    print("=" * 60)


def main():
    print("=" * 60)
    print("小红书批量博主爬虫")
    print("=" * 60)
    print()

    parser = argparse.ArgumentParser(description='小红书批量博主爬虫')
    parser.add_argument('-f', '--file', type=str, default='resources/urls.txt', help='博主URL文件路径，默认resources/urls.txt')
    parser.add_argument('-n', '--num', type=int, default=50, help='每个博主采集的达标笔记数量，默认50篇')
    parser.add_argument('-l', '--likes', type=int, default=200, help='点赞数阈值，默认200')
    parser.add_argument('-c', '--check', type=int, default=40, help='每个博主只检查前N篇笔记（用于监控同行选题），默认40篇')
    parser.add_argument('--restart', action='store_true', help='忽略进度文件，从头开始')
    parser.add_argument('--gap', type=int, default=10, help='博主间间隔秒数，默认10秒')
    args = parser.parse_args()

    urls = load_urls(args.file)
    if not urls:
        print(f"URL文件为空，请检查 {args.file}")
        return

    target_count = args.num if args.num > 0 else 50

    progress = load_progress()
    if args.restart:
        clear_progress()
        progress = {'completed_urls': [], 'all_results': []}

    completed_set = set(progress['completed_urls'])
    all_results = progress['all_results']

    remaining = [url for url in urls if url not in completed_set]

    if completed_set:
        print(f"发现进度文件: 已完成 {len(completed_set)}/{len(urls)} 个博主")
        if remaining:
            print(f"剩余待处理: {len(remaining)} 个博主")
        else:
            print("所有博主已爬取完毕，直接生成汇总报告")
            if all_results:
                generate_summary_report(all_results, target_count)
            return
    else:
        print(f"从 {args.file} 加载了 {len(urls)} 个博主URL")
        for i, url in enumerate(urls, 1):
            print(f"  {i}. {url}")

    print()
    print(f"每个博主目标: {target_count} 篇点赞>{args.likes}的笔记")
    print(f"监控模式: 检查前{args.check}篇笔记")
    print(f"博主间隔: {args.gap}秒")
    print("注意: 请确保已登录小红书账号")
    print()

    spider = XHSSpider()
    spider.like_threshold = args.likes
    results = []
    all_qualified_notes = []

    try:
        for i, url in enumerate(remaining, 1):
            print(f"\n{'=' * 60}")
            print(f"正在处理第 {i}/{len(remaining)} 个博主")
            print(f"{'=' * 60}")

            if i > 1:
                gap = args.gap + random.uniform(-2, 5)
                print(f"博主间隔等待 {gap:.1f} 秒...")
                time.sleep(gap)

            spider.notes_data = []

            crawl_success = False

            try:
                result = crawl_single_blogger_check(spider, url, args.check, args.likes, i, len(remaining))
                results.append(result)
                crawl_success = result['success']
                
                if result.get('qualified_notes'):
                    all_qualified_notes.extend(result['qualified_notes'])

            except Exception as e:
                print(f"博主爬取出错: {e}")
                import traceback
                traceback.print_exc()
                results.append({
                    'url': url,
                    'name': '错误',
                    'total': 0,
                    'qualified': 0,
                    'qualified_notes': [],
                    'success': False,
                })

            if url not in completed_set:
                progress['completed_urls'].append(url)
                progress['all_results'] = results
                save_progress(progress)
                print(f"进度已保存 (已完成 {len(progress['completed_urls'])}/{len(urls)})")

        print(f"\n{'=' * 60}")
        print("所有博主爬取完毕")
        print(f"{'=' * 60}")

        print(f"\n各博主统计:")
        total_all = 0
        qualified_all = 0
        for r in results:
            status = "✓" if r['success'] else "✗"
            print(f"  [{status}] {r['name'][:15]:<15} | 总计: {r['total']:>4} | 达标: {r['qualified']:>4}")
            total_all += r['total']
            qualified_all += r['qualified']
        print(f"  {'─' * 45}")
        print(f"  {'合计':<15} | 总计: {total_all:>4} | 达标: {qualified_all:>4}")

        generate_summary_report(results, target_count, check_mode=True, check_count=args.check)

        excel_file = None
        if all_qualified_notes:
            excel_file = save_monitor_excel(all_qualified_notes, args.likes)

        print()
        print("=" * 60)
        print("批量爬取完成！")
        print(f"博主数: {len(urls)}")
        print(f"总笔记数: {total_all}")
        print(f"总达标数: {qualified_all}")
        bloggers_with_qualified = sum(1 for r in results if r['qualified'] > 0)
        print(f"有达标笔记的博主: {bloggers_with_qualified}/{len(urls)}")
        if excel_file:
            print(f"监控汇总文件: {excel_file}")
        print("=" * 60)

        clear_progress()

        spider.close()

    except KeyboardInterrupt:
        print(f"\n\n用户中断！进度已保存，下次运行将从断点继续")
        print(f"已完成: {len(progress['completed_urls'])}/{len(urls)} 个博主")
        spider.close()

    except Exception as e:
        print(f"批量爬取过程中出错: {e}")
        import traceback
        traceback.print_exc()
        print(f"进度已保存，下次运行将从断点继续")
        input("\n按回车键关闭浏览器...")
        spider.close()


if __name__ == '__main__':
    main()
