from DrissionPage import ChromiumPage
import json
import time
import re
import os
from urllib.parse import urlencode, parse_qs, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class XHSSpider:
    def __init__(self):
        self.page = ChromiumPage()
        self.notes_data = []
        self.user_name = ''

    def parse_user_id_from_url(self, url):
        """从小红书主页URL中提取用户ID"""
        patterns = [
            r'/user/profile/(\w+)',
            r'/user/profile/([^?/]+)',
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return None

    def get_user_name(self):
        """获取博主用户名"""
        try:
            name_selectors = [
                'css:.user-name',
                'css:.user-nickname',
                'css:[class*="userName"]',
                'css:[class*="nickname"]',
                'css:.user-info .name',
                'css:.profile-name',
            ]
            for selector in name_selectors:
                try:
                    name_elem = self.page.ele(selector, timeout=1)
                    if name_elem:
                        name = name_elem.text.strip()
                        if name:
                            self.user_name = name
                            print(f"获取到博主用户名: {name}")
                            return name
                except:
                    continue

            js_code = """
            const nameElem = document.querySelector('.user-name, .user-nickname, [class*="userName"], [class*="nickname"]');
            return nameElem ? nameElem.innerText.trim() : '';
            """
            name = self.page.run_js(js_code)
            if name:
                self.user_name = name
                print(f"获取到博主用户名: {name}")
                return name
        except Exception as e:
            print(f"获取用户名失败: {e}")

        self.user_name = 'unknown_user'
        return self.user_name

    def scroll_page(self, scroll_times=20, scroll_pause=3):
        """滚动页面加载更多笔记"""
        print(f"开始滚动页面，计划滚动 {scroll_times} 次...")
        for i in range(scroll_times):
            self.page.scroll.down(400)
            time.sleep(scroll_pause)
            print(f"第 {i + 1}/{scroll_times} 次滚动完成")

    def _extract_notes_from_page(self):
        """从当前页面提取笔记"""
        note_selectors = [
            'css:.note-item',
            'css:[class*="note-item"]',
            'css:.feeds-page .note-item',
            'css:.user-page .note-item',
            'css:section[class*="note"] > div',
            'css:.waterfall-item',
        ]

        for selector in note_selectors:
            try:
                elements = self.page.eles(selector, timeout=2)
                if elements and len(elements) > 0:
                    for note_elem in elements:
                        try:
                            note_info = self.extract_note_info(note_elem)
                            if note_info and note_info['note_id']:
                                if not any(n['note_id'] == note_info['note_id'] for n in self.notes_data):
                                    self.notes_data.append(note_info)
                        except:
                            continue
                    break
            except:
                continue

    def scroll_and_extract(self, scroll_times=20, scroll_pause=3):
        """滚动页面并逐步提取笔记"""
        print(f"开始滚动并提取，计划滚动 {scroll_times} 次...")
        for i in range(scroll_times):
            self.page.scroll.down(400)
            time.sleep(scroll_pause)

            # 每滚动2次提取一次笔记
            if (i + 1) % 2 == 0:
                self._extract_notes_from_page()
                print(f"第 {i + 1}/{scroll_times} 次滚动完成，当前已提取 {len(self.notes_data)} 篇笔记")
            else:
                print(f"第 {i + 1}/{scroll_times} 次滚动完成")

    def extract_note_info(self, note_element):
        """从笔记元素中提取信息"""
        try:
            # 获取笔记链接
            link_elem = note_element.ele('css:a[href*="/explore/"]', timeout=0.5)
            if not link_elem:
                return None

            note_url = link_elem.attr('href')
            if not note_url.startswith('http'):
                note_url = 'https://www.xiaohongshu.com' + note_url

            # 获取笔记ID
            note_id_match = re.search(r'/explore/(\w+)', note_url)
            note_id = note_id_match.group(1) if note_id_match else ''

            # 获取标题
            title = ''
            try:
                title_elem = note_element.ele('css:.title, .desc, span[class*="title"]', timeout=0.5)
                if title_elem:
                    title = title_elem.text.strip()
            except:
                pass

            # 如果没有标题，尝试获取描述
            if not title:
                try:
                    desc_elem = note_element.ele('css:.desc span, .content span', timeout=0.5)
                    if desc_elem:
                        title = desc_elem.text.strip()[:50]  # 截取前50字
                except:
                    pass

            # 获取点赞数
            likes = '0'
            try:
                # 尝试多种点赞数选择器
                like_selectors = [
                    'css:.like-wrapper .count',
                    'css:.likes .count',
                    'css:span[class*="like"]',
                    'css:.interaction span',
                    'css:.count',
                    'css:[class*="like"] span',
                    'css:.info span',
                    'css:.meta span',
                    'css:span',
                ]
                for selector in like_selectors:
                    try:
                        like_elem = note_element.ele(selector, timeout=0.3)
                        if like_elem:
                            likes_text = like_elem.text.strip()
                            # 检查是否为数字格式（包括万、w、k等）
                            if likes_text and any(c.isdigit() for c in likes_text):
                                likes = likes_text
                                break
                    except:
                        continue
            except Exception as e:
                print(f"获取点赞数出错: {e}")
                pass

            # 获取封面图
            cover_image = ''
            try:
                img_elem = note_element.ele('css:img[class*="img"], css:.cover img', timeout=0.3)
                if img_elem:
                    cover_image = img_elem.attr('src') or img_elem.attr('data-src')
            except:
                pass

            return {
                'note_id': note_id,
                'title': title,
                'likes': likes,
                'note_url': note_url,
                'cover_image': cover_image,
            }
        except Exception as e:
            print(f"提取笔记信息时出错: {e}")
            return None

    def count_qualified_notes(self):
        """统计点赞>90的笔记数量"""
        return sum(1 for note in self.notes_data if self.parse_likes(note.get('likes', '0')) > 90)

    def crawl_user_notes(self, user_url, target_count=50):
        """爬取用户主页的笔记，直到达到目标数量"""
        print(f"正在访问用户主页: {user_url}")

        self.page.get(user_url)
        
        import random

        # 模拟人类浏览行为 - 页面加载后的随机等待
        page_load_delay = random.uniform(2, 4)
        print(f"等待页面加载中... ({page_load_delay:.1f}秒)")
        time.sleep(page_load_delay)

        # 随机延迟（模拟人类查看页面）
        if random.random() < 0.5:
            for _ in range(random.randint(2, 5)):
                time.sleep(random.uniform(0.1, 0.3))

        self.get_user_name()

        # 先提取页面已有的笔记（前几条）
        print("正在提取初始笔记...")
        self._extract_notes_from_page()

        qualified_count = self.count_qualified_notes()
        print(f"当前达标笔记数: {qualified_count}/{target_count}")

        # 如果还没达到目标，继续滚动提取
        max_scrolls = 50  # 最大滚动次数，防止无限循环
        no_new_count = 0  # 连续没有新笔记的次数

        import random

        for scroll_idx in range(max_scrolls):
            if qualified_count >= target_count:
                print(f"\n已达到目标数量 {target_count} 篇达标笔记，停止爬取")
                break

            # 模拟人类滚动行为 - 随机滚动距离和速度
            scroll_distance = random.randint(300, 600)  # 随机滚动距离
            scroll_pause = random.uniform(1.5, 3.5)  # 随机暂停时间
            
            # 滚动页面
            self.page.scroll.down(scroll_distance)
            
            # 随机添加一些人类行为
            if random.random() < 0.3:  # 30%概率
                # 随机延迟（模拟人类查看页面）
                time.sleep(random.uniform(0.2, 0.5))

            time.sleep(scroll_pause)

            # 提取笔记
            prev_count = len(self.notes_data)
            self._extract_notes_from_page()
            new_notes = len(self.notes_data) - prev_count

            qualified_count = self.count_qualified_notes()

            if new_notes == 0:
                no_new_count += 1
                print(f"第 {scroll_idx + 1} 次滚动: 无新笔记 (连续{no_new_count}次)")
                # 连续3次没有新笔记，认为已经到底
                if no_new_count >= 3:
                    print("\n已连续3次没有新笔记，认为已到达页面底部")
                    break
            else:
                no_new_count = 0
                print(f"第 {scroll_idx + 1} 次滚动: 新增{new_notes}篇，达标{qualified_count}/{target_count}")

        if qualified_count < target_count:
            print(f"\n博主笔记已爬取完毕，达标笔记仅{qualified_count}篇（目标{target_count}篇）")

        print(f"\n共提取到 {len(self.notes_data)} 篇笔记，其中达标{qualified_count}篇")

        # 打印所有笔记的点赞数（用于调试）
        print("\n提取到的笔记数据预览（前10篇）：")
        for i, note in enumerate(self.notes_data[:10], 1):
            parsed_likes = self.parse_likes(note['likes'])
            qualified = "✓" if parsed_likes > 90 else "✗"
            print(f"  {i}. [{qualified}] 点赞: {note['likes']:>8} -> {parsed_likes:>6} | {note['title'][:25]}...")
        if len(self.notes_data) > 10:
            print(f"  ... 还有 {len(self.notes_data) - 10} 篇笔记")

        return self.notes_data

    def parse_likes(self, likes_str):
        """解析点赞数为数字"""
        if not likes_str:
            return 0
        likes_str = str(likes_str).strip()
        try:
            if '万' in likes_str:
                num = float(likes_str.replace('万', ''))
                return int(num * 10000)
            elif 'w' in likes_str.lower():
                num = float(likes_str.lower().replace('w', ''))
                return int(num * 10000)
            elif 'k' in likes_str.lower():
                num = float(likes_str.lower().replace('k', ''))
                return int(num * 1000)
            else:
                return int(float(likes_str))
        except:
            return 0

    def save_to_excel(self):
        """保存数据到Excel文件"""
        filtered_notes = [
            note for note in self.notes_data
            if self.parse_likes(note.get('likes', '0')) > 90
        ]

        print(f"原始笔记数: {len(self.notes_data)}, 筛选后(点赞>90): {len(filtered_notes)}")

        if not filtered_notes:
            print("没有点赞数大于90的笔记，不生成Excel文件")
            return None

        # 创建data目录（如果不存在）
        data_dir = 'data'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
            print(f"创建目录: {data_dir}/")

        filename = os.path.join(data_dir, f"{self.user_name}_notes.xlsx" if self.user_name else 'xhs_notes.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = "笔记数据"

        headers = ['序号', '标题', '点赞数', '详情页URL']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        link_font = Font(color='0563C1', underline='single')

        for row, note in enumerate(filtered_notes, 2):
            ws.cell(row=row, column=1, value=row - 1).border = thin_border
            ws.cell(row=row, column=2, value=note.get('title', '')).border = thin_border
            ws.cell(row=row, column=3, value=note.get('likes', '0')).border = thin_border

            url_cell = ws.cell(row=row, column=4)
            url_value = note.get('note_url', '')
            url_cell.value = url_value
            url_cell.hyperlink = url_value
            url_cell.font = link_font
            url_cell.border = thin_border

            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 60

        ws.freeze_panes = 'A2'

        wb.save(filename)
        print(f"数据已保存到: {filename}")
        return filename

    def close(self):
        """关闭浏览器"""
        self.page.quit()
        print("浏览器已关闭")


def load_urls_from_file(filename='resources/urls.txt'):
    """从文件加载URL列表"""
    urls = []
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    urls.append(line)
    except FileNotFoundError:
        print(f"文件 {filename} 不存在")
    return urls


def save_progress(progress_file, completed_urls):
    """保存进度到文件"""
    with open(progress_file, 'w', encoding='utf-8') as f:
        for url in completed_urls:
            f.write(url + '\n')


def load_progress(progress_file):
    """加载已完成的URL列表"""
    completed = set()
    try:
        with open(progress_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    completed.add(line)
    except FileNotFoundError:
        pass
    return completed


def crawl_single_blogger(spider, url, target_count, blogger_index, total_count):
    """爬取单个博主"""
    print("\n" + "=" * 60)
    print(f"正在处理第 {blogger_index}/{total_count} 个博主")
    print(f"URL: {url}")
    print("=" * 60)

    try:
        notes = spider.crawl_user_notes(url, target_count=target_count)

        if notes:
            excel_file = spider.save_to_excel()
            qualified_count = spider.count_qualified_notes()

            print()
            print("-" * 60)
            print(f"博主 {spider.user_name} 处理完成")
            print(f"总笔记: {len(notes)} 篇，达标: {qualified_count} 篇")
            if excel_file:
                print(f"保存文件: {excel_file}")
            print("-" * 60)

            return {
                'url': url,
                'name': spider.user_name,
                'total': len(notes),
                'qualified': qualified_count,
                'file': excel_file,
                'success': True
            }
        else:
            print(f"未从该博主获取到笔记数据")
            return {
                'url': url,
                'name': '未知',
                'total': 0,
                'qualified': 0,
                'file': None,
                'success': False
            }

    except Exception as e:
        print(f"处理博主时出错: {e}")
        import traceback
        traceback.print_exc()
        return {
            'url': url,
            'name': '错误',
            'total': 0,
            'qualified': 0,
            'file': None,
            'success': False
        }


def generate_summary_report(results, target_count):
    """生成汇总报告"""
    print("\n" + "=" * 60)
    print("批量爬取汇总报告")
    print("=" * 60)

    total_bloggers = len(results)
    success_bloggers = sum(1 for r in results if r['success'])
    total_notes = sum(r['total'] for r in results)
    total_qualified = sum(r['qualified'] for r in results)

    print(f"\n总体统计:")
    print(f"  博主总数: {total_bloggers}")
    print(f"  成功爬取: {success_bloggers}")
    print(f"  失败/无数据: {total_bloggers - success_bloggers}")
    print(f"  总笔记数: {total_notes}")
    print(f"  总达标数(点赞>{target_count}): {total_qualified}")

    print(f"\n各博主详情:")
    for i, result in enumerate(results, 1):
        status = "✓" if result['success'] else "✗"
        print(f"  {i}. [{status}] {result['name'][:15]:<15} | 达标: {result['qualified']:>3}/{target_count} | 总笔记: {result['total']:>3}")

    # 保存汇总报告到文本文件（data目录）
    data_dir = 'data'
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)

    report_file = os.path.join(data_dir, f'batch_report_{time.strftime("%Y%m%d_%H%M%S")}.txt')
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("批量爬取汇总报告\n")
        f.write("=" * 60 + "\n\n")
        f.write(f"总体统计:\n")
        f.write(f"  博主总数: {total_bloggers}\n")
        f.write(f"  成功爬取: {success_bloggers}\n")
        f.write(f"  失败/无数据: {total_bloggers - success_bloggers}\n")
        f.write(f"  总笔记数: {total_notes}\n")
        f.write(f"  总达标数(点赞>90): {total_qualified}\n\n")
        f.write("各博主详情:\n")
        for i, result in enumerate(results, 1):
            status = "成功" if result['success'] else "失败"
            f.write(f"  {i}. [{status}] {result['name']}\n")
            f.write(f"      URL: {result['url']}\n")
            f.write(f"      达标: {result['qualified']}/{target_count}, 总笔记: {result['total']}\n")
            if result['file']:
                f.write(f"      文件: {result['file']}\n")
            f.write("\n")

    print(f"\n汇总报告已保存到: {report_file}")
    print("=" * 60)


def main():
    print("=" * 60)
    print("小红书博主笔记爬虫 - 批量模式")
    print("=" * 60)
    print()

    # 询问模式
    print("请选择运行模式:")
    print("1. 单博主模式")
    print("2. 批量模式(从resources/urls.txt读取)")
    mode = input("请输入选项(1或2，默认1): ").strip() or "1"

    if mode == "1":
        run_single_mode()
    else:
        run_batch_mode()


def run_single_mode():
    """单博主模式"""
    print("\n--- 单博主模式 ---\n")

    user_url = input("请输入小红书博主主页URL: ").strip()

    if not user_url:
        print("URL不能为空！")
        return

    if 'xiaohongshu.com' not in user_url:
        print("请输入有效的小红书URL！")
        return

    target_input = input("请输入需要采集的达标笔记数量(点赞>90，默认50篇): ").strip()
    target_count = int(target_input) if target_input.isdigit() else 50

    print()
    print("正在启动爬虫...")
    print(f"目标: 采集 {target_count} 篇点赞>90的笔记")
    print("注意: 请确保已登录小红书账号，否则可能无法获取完整数据")
    print()

    spider = XHSSpider()

    try:
        notes = spider.crawl_user_notes(user_url, target_count=target_count)

        if notes:
            excel_file = spider.save_to_excel()

            print()
            print("=" * 60)
            print("爬取完成！")
            print(f"博主: {spider.user_name}")
            print(f"共获取 {len(notes)} 篇笔记")
            print(f"数据已保存到: {excel_file}")
            print("=" * 60)
        else:
            print("未获取到任何笔记数据")

    except Exception as e:
        print(f"爬取过程中出错: {e}")
        import traceback
        traceback.print_exc()

    finally:
        input("\n按回车键关闭浏览器...")
        spider.close()


def run_batch_mode():
    """批量模式"""
    print("\n--- 批量模式 ---\n")

    # 加载URL列表
    urls = load_urls_from_file('resources/urls.txt')

    if not urls:
        print("resources/urls.txt 文件为空或不存在，请先添加博主URL")
        print("格式: 每行一个URL，以#开头的行为注释")
        return

    print(f"从 resources/urls.txt 加载了 {len(urls)} 个博主URL")

    # 设置目标数量
    target_input = input("请输入每个博主需要采集的达标笔记数量(点赞>90，默认30篇): ").strip()
    target_count = int(target_input) if target_input.isdigit() else 30

    # 加载进度
    progress_file = 'progress.txt'
    completed_urls = load_progress(progress_file)

    if completed_urls:
        print(f"发现已完成的进度，共 {len(completed_urls)} 个博主")
        skip_completed = input("是否跳过已完成的博主?(y/n，默认y): ").strip().lower() != 'n'
    else:
        skip_completed = False

    # 过滤已完成的URL
    if skip_completed:
        urls = [url for url in urls if url not in completed_urls]
        print(f"剩余待处理博主: {len(urls)} 个")

    if not urls:
        print("所有博主已处理完毕！")
        return

    print()
    print("正在启动批量爬虫...")
    print(f"每个博主目标: {target_count} 篇点赞>90的笔记")
    print("注意: 请确保已登录小红书账号")
    print()

    # 创建爬虫实例
    spider = XHSSpider()
    results = []

    try:
        for i, url in enumerate(urls, 1):
            result = crawl_single_blogger(spider, url, target_count, i, len(urls))
            results.append(result)

            # 保存进度
            if result['success']:
                completed_urls.add(url)
                save_progress(progress_file, completed_urls)

            # 询问是否继续（每完成3个博主询问一次）
            if i < len(urls) and i % 3 == 0:
                cont = input(f"\n已完成 {i}/{len(urls)} 个博主，是否继续?(y/n，默认y): ").strip().lower()
                if cont == 'n':
                    print("用户中断批量处理")
                    break

        # 生成汇总报告
        generate_summary_report(results, target_count)

    except Exception as e:
        print(f"批量处理过程中出错: {e}")
        import traceback
        traceback.print_exc()

    finally:
        input("\n按回车键关闭浏览器...")
        spider.close()


if __name__ == '__main__':
    main()
