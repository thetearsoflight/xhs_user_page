from DrissionPage import ChromiumPage
import json
import time
import random
import re
import os
import argparse
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class XHSKeywordSpiderV2:
    API_KEYWORD = 'web/v1/search/notes'

    def __init__(self):
        self.page = ChromiumPage()
        self.notes_data = []
        self.seen_ids = set()
        self.keyword = ''
        self.like_threshold = 90

    def search_by_keyword(self, keyword, sort_by_time=True):
        self.keyword = keyword
        encoded_keyword = quote(keyword)
        search_url = f"https://www.xiaohongshu.com/search_result?keyword={encoded_keyword}&source=web_explore_feed"
        if sort_by_time:
            search_url += "&sort=general_note_type1"
        print(f"正在搜索关键词: {keyword}")
        print(f"访问URL: {search_url}")

        self.page.listen.start(self.API_KEYWORD)
        self.page.get(search_url)

        page_load_delay = random.uniform(2, 4)
        print(f"等待页面加载中... ({page_load_delay:.1f}秒)")
        time.sleep(page_load_delay)

        if sort_by_time:
            self._try_click_latest_filter()

    def _try_click_latest_filter(self):
        print("尝试点击'最新'筛选...")
        try:
            shaixuan = self.page.ele('xpath://span[text()="筛选"]', timeout=5)
            if shaixuan:
                shaixuan.hover()
                time.sleep(2)
                latest_btn = self.page.ele('xpath://span[text()="最新"]', timeout=3)
                if latest_btn:
                    latest_btn.click()
                    print("已点击'最新'选项")
                    time.sleep(3)
                else:
                    print("未找到'最新'选项，可能URL参数已生效")
            else:
                print("未找到'筛选'按钮，可能URL参数已生效")
        except Exception as e:
            print(f"筛选操作失败({e})，继续使用当前排序")

    def _parse_note_from_api(self, note_item):
        note_id = note_item.get('id', '')
        if not note_id or note_id in self.seen_ids:
            return None

        note_card = note_item.get('note_card', note_item)
        title = (note_card.get('display_title', '')
                 or note_card.get('title', '')
                 or note_card.get('displayTitle', ''))

        interact_info = note_card.get('interact_info', note_card.get('interactInfo', {}))
        liked_count_str = (interact_info.get('liked_count', '')
                           or interact_info.get('likedCount', '0'))

        user_info = note_card.get('user', {})
        nickname = user_info.get('nickname', '') or user_info.get('nick_name', '')

        note_type = note_card.get('type', '')

        note_url = f"https://www.xiaohongshu.com/explore/{note_id}"

        cover = note_card.get('cover', {})
        cover_url = ''
        if isinstance(cover, dict):
            cover_url = cover.get('url_default', '') or cover.get('url_pre', '') or cover.get('url', '')

        likes_num = self.parse_likes(liked_count_str)

        self.seen_ids.add(note_id)

        return {
            'note_id': note_id,
            'title': title,
            'likes': liked_count_str,
            'likes_num': likes_num,
            'nickname': nickname,
            'note_type': note_type,
            'note_url': note_url,
            'cover_image': cover_url,
        }

    def _process_api_response(self, packet, debug=False):
        try:
            body = packet.response.body
            if not body or not isinstance(body, dict):
                print(f"[DEBUG] body类型: {type(body)}, 内容前200字: {str(body)[:200]}")
                return 0

            data = body.get('data', {})
            items = data.get('items', [])
            if not items:
                has_more = data.get('has_more', False)
                if not has_more:
                    return -1
                return 0

            if debug:
                print(f"\n[DEBUG] 第一条笔记完整JSON结构:")
                print(json.dumps(items[0], ensure_ascii=False, indent=2)[:2000])
                print(f"[DEBUG] items数量: {len(items)}")

            new_count = 0
            for item in items:
                note_info = self._parse_note_from_api(item)
                if note_info:
                    self.notes_data.append(note_info)
                    new_count += 1

            return new_count
        except Exception as e:
            print(f"解析API响应出错: {e}")
            import traceback
            traceback.print_exc()
            return 0

    def crawl_keyword_notes(self, keyword, target_count=50, sort_by_time=True):
        self.search_by_keyword(keyword, sort_by_time)

        print("等待搜索接口返回数据...")
        packet = self.page.listen.wait(timeout=15)
        if packet:
            self._process_api_response(packet, debug=True)
            print(f"初始加载: 获取到 {len(self.notes_data)} 篇笔记")

        qualified_count = self.count_qualified_notes()
        print(f"当前达标笔记数: {qualified_count}/{target_count}")

        max_scrolls = 80
        no_new_count = 0
        max_no_new = 5

        for scroll_idx in range(max_scrolls):
            if qualified_count >= target_count:
                print(f"\n已达到目标数量 {target_count} 篇达标笔记，停止爬取")
                break

            scroll_distance = random.randint(300, 600)
            self.page.scroll.down(scroll_distance)

            if random.random() < 0.3:
                time.sleep(random.uniform(0.2, 0.5))

            scroll_pause = random.uniform(1.5, 3.5)
            time.sleep(scroll_pause)

            packet = self.page.listen.wait(timeout=10)
            if not packet:
                no_new_count += 1
                print(f"第 {scroll_idx + 1} 次滚动: 未捕获到接口响应 (连续{no_new_count}次)")
                if no_new_count >= max_no_new:
                    print(f"\n已连续{max_no_new}次未捕获到新数据，认为已到达末尾")
                    break
                continue

            prev_total = len(self.notes_data)
            result = self._process_api_response(packet)

            if result == -1:
                print("接口返回has_more=false，已到达末尾")
                break

            new_notes = len(self.notes_data) - prev_total
            qualified_count = self.count_qualified_notes()

            if new_notes == 0:
                no_new_count += 1
                print(f"第 {scroll_idx + 1} 次滚动: 无新笔记 (连续{no_new_count}次)")
                if no_new_count >= max_no_new:
                    print(f"\n已连续{max_no_new}次无新笔记，认为已到达末尾")
                    break
            else:
                no_new_count = 0
                print(f"第 {scroll_idx + 1} 次滚动: 新增{new_notes}篇，达标{qualified_count}/{target_count}")

        self.page.listen.stop()

        if qualified_count < target_count:
            print(f"\n搜索结果已爬取完毕，达标笔记仅{qualified_count}篇（目标{target_count}篇）")

        print(f"\n共提取到 {len(self.notes_data)} 篇笔记，其中达标{qualified_count}篇")

        self._print_preview()

        return self.notes_data

    def _print_preview(self):
        print("\n提取到的笔记数据预览（前10篇）：")
        for i, note in enumerate(self.notes_data[:10], 1):
            qualified = "✓" if note['likes_num'] > self.like_threshold else "✗"
            print(f"  {i}. [{qualified}] 点赞: {note['likes']:>8} -> {note['likes_num']:>6} | {note['title'][:30]}")
        if len(self.notes_data) > 10:
            print(f"  ... 还有 {len(self.notes_data) - 10} 篇笔记")

    def count_qualified_notes(self):
        return sum(1 for note in self.notes_data if note.get('likes_num', 0) > self.like_threshold)

    def parse_likes(self, likes_str):
        if not likes_str:
            return 0
        likes_str = str(likes_str).strip()
        try:
            if '万' in likes_str:
                num = float(likes_str.replace('万', ''))
                return int(num * 10000)
            elif 'w' in likes_str.lower():
                num = float(likes_str.lower().replace('w', ''))
                return int(num * 10000)
            elif 'k' in likes_str.lower():
                num = float(likes_str.lower().replace('k', ''))
                return int(num * 1000)
            else:
                return int(float(likes_str))
        except (ValueError, TypeError):
            return 0

    def save_to_excel(self):
        filtered_notes = [
            note for note in self.notes_data
            if note.get('likes_num', 0) > self.like_threshold
        ]

        print(f"原始笔记数: {len(self.notes_data)}, 筛选后(点赞>{self.like_threshold}): {len(filtered_notes)}")

        if not filtered_notes:
            print("没有点赞数大于阈值的笔记，不生成Excel文件")
            return None

        data_dir = 'data'
        os.makedirs(data_dir, exist_ok=True)

        filename = os.path.join(data_dir, f"{self.keyword}_notes.xlsx" if self.keyword else 'xhs_notes.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = "笔记数据"

        headers = ['序号', '标题', '作者', '类型', '点赞数', '详情页URL']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        link_font = Font(color='0563C1', underline='single')

        for row, note in enumerate(filtered_notes, 2):
            ws.cell(row=row, column=1, value=row - 1).border = thin_border
            ws.cell(row=row, column=2, value=note.get('title', '')).border = thin_border
            ws.cell(row=row, column=3, value=note.get('nickname', '')).border = thin_border

            note_type = note.get('note_type', '')
            type_display = '视频' if note_type == 'video' else '图文'
            ws.cell(row=row, column=4, value=type_display).border = thin_border

            ws.cell(row=row, column=5, value=note.get('likes', '0')).border = thin_border

            url_cell = ws.cell(row=row, column=6)
            url_value = note.get('note_url', '')
            url_cell.value = url_value
            url_cell.hyperlink = url_value
            url_cell.font = link_font
            url_cell.border = thin_border

            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 55

        ws.freeze_panes = 'A2'

        wb.save(filename)
        print(f"数据已保存到: {filename}")
        return filename

    def close(self):
        try:
            self.page.listen.stop()
        except Exception:
            pass
        self.page.quit()
        print("浏览器已关闭")


def main():
    print("=" * 60)
    print("小红书关键词搜索爬虫 V2（接口监听版）")
    print("=" * 60)
    print()

    parser = argparse.ArgumentParser(description='小红书关键词搜索爬虫V2')
    parser.add_argument('-k', '--keyword', type=str, help='搜索关键词')
    parser.add_argument('-n', '--num', type=int, default=50, help='需要采集的达标笔记数量(点赞>200)，默认50篇')
    parser.add_argument('-l', '--likes', type=int, default=200, help='点赞数阈值，默认200')
    parser.add_argument('--no-sort-time', action='store_true', help='不按最新排序（默认按最新排序）')
    args = parser.parse_args()

    if args.keyword:
        keyword = args.keyword.strip()
    else:
        keyword = input("请输入搜索关键词: ").strip()

    if not keyword:
        print("关键词不能为空！")
        return

    target_count = args.num if args.num > 0 else 50
    sort_by_time = not args.no_sort_time

    print()
    print("正在启动爬虫...")
    print(f"关键词: {keyword}")
    print(f"目标: 采集 {target_count} 篇点赞>{args.likes}的笔记")
    print(f"排序: {'最新' if sort_by_time else '综合'}")
    print("注意: 请确保已登录小红书账号，否则可能无法获取完整数据")
    print()

    spider = XHSKeywordSpiderV2()
    spider.like_threshold = args.likes

    try:
        notes = spider.crawl_keyword_notes(keyword, target_count=target_count, sort_by_time=sort_by_time)

        if notes:
            excel_file = spider.save_to_excel()

            print()
            print("=" * 60)
            print("爬取完成！")
            print(f"关键词: {keyword}")
            print(f"共获取 {len(notes)} 篇笔记")
            qualified = spider.count_qualified_notes()
            print(f"达标(点赞>{spider.like_threshold}): {qualified} 篇")
            if excel_file:
                print(f"Excel数据已保存到: {excel_file}")
            print("=" * 60)
        else:
            print("未获取到任何笔记数据")
            print("可能原因：")
            print("  1. 未登录小红书账号")
            print("  2. 搜索接口URL已变更（请检查API_KEYWORD常量）")
            print("  3. 网络问题")

        spider.close()

    except Exception as e:
        print(f"爬取过程中出错: {e}")
        import traceback
        traceback.print_exc()
        input("\n按回车键关闭浏览器...")
        spider.close()


if __name__ == '__main__':
    main()
